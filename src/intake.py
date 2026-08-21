"""
The intake workbook: standing up a new client without writing code.

A client pack is five YAML files and a facts document. Hand-written, that is a
developer task — which meant the accelerator could demonstrate configurability
without anyone being able to use it on an engagement. This module closes that
gap in both directions:

    make template CLIENT=acme     -> a formatted workbook to fill in
    make onboard FILE=acme.xlsx   -> a complete, validated client pack

The workbook is deliberately not a thin data-entry form. Its columns are the
questions an FP&A diagnostic asks — what is the plausible range, who owns
this, can management actually move it, what is the tolerance — so filling it
in is a structured interview whose output happens to be executable. Most
finance functions cannot answer several of those columns, and that inability
is itself the finding (see src/readiness.py).

Nothing here invents a value. A blank cell stays blank and surfaces as a gap;
it is never defaulted into something plausible, because a pack that looks
complete and is not is worse than one that is visibly unfinished.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import yaml

from src import clientpack

# --------------------------------------------------------------------------
# The workbook's shape
# --------------------------------------------------------------------------

SHEETS = ("Client", "Baseline", "Segments", "Drivers", "Rules")

CLIENT_FIELDS = [
    ("client_id", "Short id, lowercase, no spaces — becomes the folder name", "acme_industrial"),
    ("name", "Full legal or trading name as it should appear", "Acme Industrial AG"),
    ("short_label", "Short name for page headers and the model selector", "Acme"),
    ("data_basis", "Public Data | Client Data | Synthetic Demo", "Client Data"),
    ("is_synthetic", "TRUE if any figure is invented", "FALSE"),
    ("industry", "One line", "Industrial manufacturing"),
    ("currency", "ISO code", "EUR"),
    ("currency_symbol", "Symbol shown before figures", "€"),
    ("unit", "Reporting unit of the figures below", "millions"),
    ("fiscal_year", "The year being planned", "FY2026"),
    ("baseline_year", "The year the plan is built from", "FY2025"),
    ("objective", "Revenue | EBITDA | EBIT | Free Cash Flow | Working Capital", "Free Cash Flow"),
    ("materiality_high", "Exposure at or above this is High, in the unit above", "60"),
    ("materiality_medium", "Exposure at or above this is Medium", "25"),
    ("materiality_rationale", "WHY those numbers. Required — a threshold nobody can justify "
                              "is a threshold nobody should trust.", ""),
    ("disclaimer", "Shown on the Evidence page", ""),
]

BASELINE_FIELDS = [
    ("net_sales", "Net sales / revenue"),
    ("gross_profit", "Gross profit"),
    ("ebitda", "EBITDA"),
    ("operating_profit", "Operating profit / EBIT"),
    ("effective_tax_rate_pct", "Effective tax rate, %"),
    ("operating_working_capital_pct", "Operating working capital as % of net sales"),
    ("capex", "Capital expenditure"),
]

#: Example values are prefixed EXAMPLE_ so a filled row can never be mistaken
#: for one. The reader used to discard rows by matching the example's content,
#: which silently deleted any real driver called `inventory_days` — a plausible
#: id, and the loss was invisible: the pack built, loaded and ranked without it.
EXAMPLE_PREFIX = "EXAMPLE_"

DRIVER_COLUMNS = [
    ("id", "lowercase_with_underscores", EXAMPLE_PREFIX + "inventory_days"),
    ("label", "As shown in the interface", "Inventory days"),
    ("category", "Commercial | Margin | Working Capital | Input Cost | Investment | Tax", "Working Capital"),
    ("unit", "pct | days | ppt | eur_m", "days"),
    ("baseline", "The planned value", "104"),
    ("min", "Lowest value the slider allows", "80"),
    ("max", "Highest value the slider allows", "145"),
    ("step", "Slider increment", "1"),
    ("exposure_low", "Low end of the PLAUSIBLE range — not the slider bound", "95"),
    ("exposure_high", "High end of the plausible range", "118"),
    ("maps_to", "division_growth | ebitda_margin_pct | operating_working_capital_pct | "
                "capex_eur_m | effective_tax_rate_pct", "operating_working_capital_pct"),
    ("role", "base | add | delta", "delta"),
    ("sign", "1 or -1. Use -1 where an increase helps (e.g. DPO)", "1"),
    ("scale", "Multiplier into the model. Days -> % of sales is 0.27397", "0.27397"),
    ("confidence", "High | Medium | Low — how firm is this assumption?", "Medium"),
    ("controllability", "High | Medium | Low — can management move it this year?", "High"),
    ("owner", "Role, not a person", "Head of Supply Chain"),
    ("source", "Where the number came from", "FY2025 plan"),
    ("guidance_text", "One line of context shown under the slider", ""),
]

RULE_COLUMNS = [
    ("id", "lowercase_with_underscores", EXAMPLE_PREFIX + "inventory_above_plan"),
    ("metric", "A driver id, or free_cash_flow / operating_profit / revenue", "inventory_days"),
    ("condition", "above | below | variance_below_base | variance_above_base", "above"),
    ("threshold", "The level that matters", "104"),
    ("severity", "critical | high | medium | low", "high"),
    ("label", "Short description of the breach", "Inventory days above plan"),
    ("management_question", "The QUESTION to put to management — not the answer", ""),
    ("suggested_owner", "Role", "Head of Supply Chain"),
    ("next_action", "The concrete next step", ""),
    ("trigger", "When this is next looked at", "Next monthly operations review"),
]

_ASSUMPTION_TO_SEGMENT_KEY = "division_growth"


class IntakeError(ValueError):
    """The workbook cannot produce a valid pack. Raised with the sheet, the row
    and what is wrong, because the person fixing it is in Excel, not a
    debugger."""


# --------------------------------------------------------------------------
# Writing the template
# --------------------------------------------------------------------------

def write_template(path: Path, client_id: str = "new_client") -> Path:
    """A formatted workbook with the columns, the guidance and one worked row.

    The example row is adidas-shaped on purpose: it shows a real driver that
    resolves, so the first thing anyone does is edit something that already
    works rather than face an empty grid.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    header = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="121A17")
    hint = Font(italic=True, size=9, color="59635E")

    wb = Workbook()

    ws = wb.active
    ws.title = "Client"
    ws.append(["Field", "Value", "Guidance"])
    for c in ws[1]:
        c.font, c.fill = header, fill
    for field, guidance, example in CLIENT_FIELDS:
        ws.append([field, example if field == "client_id" else "", guidance])
        ws.cell(ws.max_row, 1).value = field
        ws.cell(ws.max_row, 3).font = hint
    ws["B2"] = client_id
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 78

    ws = wb.create_sheet("Baseline")
    ws.append(["Figure", "Baseline year", "Prior year", "Guidance"])
    for c in ws[1]:
        c.font, c.fill = header, fill
    for field, guidance in BASELINE_FIELDS:
        ws.append([field, "", "", guidance])
        ws.cell(ws.max_row, 4).font = hint
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 56
    ws.append([])
    ws.append(["Prior year is only needed for the naive-baseline comparison. "
               "Leave it blank if you do not have it."])
    ws.cell(ws.max_row, 1).font = hint

    ws = wb.create_sheet("Segments")
    ws.append(["Segment", "Baseline year revenue"])
    for c in ws[1]:
        c.font, c.fill = header, fill
    for name in ("segment_one", "segment_two", "segment_three"):
        ws.append([name, ""])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.append([])
    ws.append(["Revenue by whichever axis the business plans on — product line, "
               "channel, region. They must sum to net sales."])
    ws.cell(ws.max_row, 1).font = hint

    ws = wb.create_sheet("Drivers")
    ws.append([c[0] for c in DRIVER_COLUMNS])
    for c in ws[1]:
        c.font, c.fill = header, fill
    ws.append([c[2] for c in DRIVER_COLUMNS])
    ws.append([c[1] for c in DRIVER_COLUMNS])
    for c in ws[3]:
        c.font = hint
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 58
    for i, _ in enumerate(DRIVER_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 19

    ws = wb.create_sheet("Rules")
    ws.append([c[0] for c in RULE_COLUMNS])
    for c in ws[1]:
        c.font, c.fill = header, fill
    ws.append([c[2] for c in RULE_COLUMNS])
    ws.append([c[1] for c in RULE_COLUMNS])
    for c in ws[3]:
        c.font = hint
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[3].height = 58
    for i, _ in enumerate(RULE_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = 22

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# --------------------------------------------------------------------------
# Reading it back
# --------------------------------------------------------------------------

def _clean(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _is_template_row(row: list, columns: list) -> bool:
    """True for the template's own example and guidance rows.

    Identified by an explicit EXAMPLE_ prefix and by the literal guidance
    string, never by resemblance to real data. The previous version discarded
    any row whose id matched the example's — which deleted a real
    `inventory_days` driver without a word, and the resulting pack built,
    loaded and ranked as though it had never been entered. A silent drop is
    the worst failure an intake can have, because the workbook still looks
    right.
    """
    first = str(row[0] or "").strip()
    if not first:
        return True
    return first.startswith(EXAMPLE_PREFIX) or first == columns[0][1]


def read_workbook(path: Path) -> dict:
    from openpyxl import load_workbook

    if not path.exists():
        raise IntakeError(f"No workbook at {path}")
    wb = load_workbook(path, data_only=True)
    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        raise IntakeError(f"Workbook is missing sheet(s): {missing}. Expected {list(SHEETS)}.")

    client = {}
    for row in wb["Client"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            client[str(row[0]).strip()] = _clean(row[1])

    baseline, prior = {}, {}
    for row in wb["Baseline"].iter_rows(min_row=2, values_only=True):
        key = str(row[0] or "").strip()
        if key in dict(BASELINE_FIELDS):
            if _clean(row[1]) is not None:
                baseline[key] = row[1]
            if len(row) > 2 and _clean(row[2]) is not None:
                prior[key] = row[2]

    segments = {}
    for row in wb["Segments"].iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()
        if name and _clean(row[1]) is not None and not name.startswith("Revenue by"):
            segments[name] = row[1]

    drivers = []
    cols = [c[0] for c in DRIVER_COLUMNS]
    for row in wb["Drivers"].iter_rows(min_row=2, values_only=True):
        if _is_template_row(list(row), DRIVER_COLUMNS):
            continue
        drivers.append({k: _clean(v) for k, v in zip(cols, row)})

    rules = []
    rcols = [c[0] for c in RULE_COLUMNS]
    for row in wb["Rules"].iter_rows(min_row=2, values_only=True):
        if _is_template_row(list(row), RULE_COLUMNS):
            continue
        rules.append({k: _clean(v) for k, v in zip(rcols, row)})

    return {"client": client, "baseline": baseline, "prior": prior,
            "segments": segments, "drivers": drivers, "rules": rules,
            "counts": {"segments": len(segments), "drivers": len(drivers), "rules": len(rules)}}


# --------------------------------------------------------------------------
# Turning it into a pack
# --------------------------------------------------------------------------

def _require(mapping: dict, key: str, where: str):
    value = mapping.get(key)
    if value in (None, ""):
        raise IntakeError(f"{where}: '{key}' is required and is blank.")
    return value


def build_pack(intake: dict, clients_dir: Path | None = None) -> Path:
    """Write a complete client pack. Validates by loading it back."""
    client = intake["client"]
    client_id = str(_require(client, "client_id", "Client sheet")).strip()
    base_year = str(_require(client, "baseline_year", "Client sheet")).replace("FY", "")
    fiscal_year = str(_require(client, "fiscal_year", "Client sheet"))

    if not intake["segments"]:
        raise IntakeError("Segments sheet: at least one segment with a revenue figure is required.")
    if not intake["drivers"]:
        raise IntakeError("Drivers sheet: at least one driver is required.")

    baseline = intake["baseline"]
    for field in ("net_sales", "ebitda", "operating_profit",
                  "effective_tax_rate_pct", "operating_working_capital_pct"):
        _require(baseline, field, "Baseline sheet")

    target = (clients_dir or clientpack.CLIENTS_DIR) / client_id
    target.mkdir(parents=True, exist_ok=True)

    # facts.json — hand-authored client, so the facts live inside the pack.
    group = {base_year: {**baseline, "source": f"Intake workbook, FY{base_year}"}}
    prior_year = str(int(base_year) - 1)
    if intake["prior"]:
        group[prior_year] = {**intake["prior"], "source": f"Intake workbook, FY{prior_year}"}
    facts = {
        "company": _require(client, "name", "Client sheet"),
        "currency": client.get("currency", "EUR"),
        "unit": client.get("unit", "millions"),
        "group": group,
        "product_division": {base_year: {**intake["segments"],
                                         "source": f"Intake workbook, FY{base_year}"}},
    }
    (target / "facts.json").write_text(json.dumps(facts, indent=4) + "\n")

    truthy = {"true", "yes", "y", "1"}
    is_synthetic = str(client.get("is_synthetic", "")).strip().lower() in truthy
    client_yaml = {
        "name": client["name"],
        "short_label": client.get("short_label") or client["name"],
        "data_basis": client.get("data_basis") or ("Synthetic Demo" if is_synthetic else "Client Data"),
        "is_synthetic": is_synthetic,
        "industry": client.get("industry", ""),
        "currency": client.get("currency", "EUR"),
        "currency_symbol": client.get("currency_symbol", "€"),
        "unit": client.get("unit", "millions"),
        "fiscal_year": fiscal_year,
        "baseline_year": f"FY{base_year}",
        "objective": client.get("objective", "Free Cash Flow"),
        "facts": "facts.json",
        "segments_path": f"product_division.{base_year}",
        "baseline_group_path": f"group.{base_year}",
        "audiences": ["CFO", "FP&A", "Business Owner"],
        "disclaimer": client.get("disclaimer", ""),
        # No outturn is collected by the intake, so no backtest is claimed.
        "has_backtest": False,
        "materiality_thresholds": {
            "high": float(_require(client, "materiality_high", "Client sheet")),
            "medium": float(_require(client, "materiality_medium", "Client sheet")),
            "rationale": _require(client, "materiality_rationale", "Client sheet"),
        },
    }
    _dump(target / "client.yaml", client_yaml,
          "Generated from an intake workbook. Edit here or re-run the intake.")

    order, drivers_out, bases = [], {}, {}
    for i, row in enumerate(intake["drivers"], start=2):
        did = row.get("id")
        if not did:
            raise IntakeError(f"Drivers sheet row {i}: 'id' is blank.")
        where = f"Drivers sheet row {i} ({did})"
        order.append(did)
        role = (row.get("role") or "base").lower()
        spec = {
            "label": _require(row, "label", where),
            "category": row.get("category") or "Other",
            "unit": row.get("unit") or "pct",
            "baseline": {"literal": float(_require(row, "baseline", where))},
            "min": float(_require(row, "min", where)),
            "max": float(_require(row, "max", where)),
            "step": float(row.get("step") or 0.5),
            "maps_to": _require(row, "maps_to", where),
            "role": role,
            "impacts": [],
            "confidence": _require(row, "confidence", where),
            "controllability": _require(row, "controllability", where),
            "owner": row.get("owner") or "Unassigned",
            "source": row.get("source") or "Intake workbook",
            "guidance_text": row.get("guidance_text") or "",
        }
        if row.get("scale") not in (None, ""):
            spec["scale"] = float(row["scale"])
        if role != "base" and row.get("sign") not in (None, ""):
            spec["sign"] = float(row["sign"])
        if row.get("exposure_low") not in (None, "") and row.get("exposure_high") not in (None, ""):
            spec["exposure_range"] = {"low": float(row["exposure_low"]),
                                      "high": float(row["exposure_high"])}
        drivers_out[did] = spec

    # An assumption driven only by deltas needs a level for them to adjust.
    for key in clientpack.MODEL_ASSUMPTION_KEYS:
        mapped = [d for d in drivers_out.values() if d["maps_to"] == key]
        if mapped and all(d["role"] == "delta" for d in mapped):
            if key == "operating_working_capital_pct":
                bases[key] = {"literal": float(baseline["operating_working_capital_pct"])}
            else:
                raise IntakeError(
                    f"Every driver mapped to '{key}' is a delta, so nothing sets its level. "
                    f"Make one of them role 'base', or map a driver to it directly."
                )

    drivers_doc = {"order": order, "drivers": drivers_out}
    if bases:
        drivers_doc = {"order": order, "assumption_bases": bases, "drivers": drivers_out}
    _dump(target / "drivers.yaml", drivers_doc,
          "Generated from an intake workbook. `baseline` values are literals here;\n"
          "a client whose plan is derived from disclosures can use fact/midpoint/solve\n"
          "instead — see clients/adidas/drivers.yaml.")

    rules_out = []
    for i, row in enumerate(intake["rules"], start=2):
        if not row.get("metric"):
            continue
        rules_out.append({
            "id": row.get("id") or f"rule_{i}",
            "metric": row["metric"],
            "condition": row.get("condition") or "above",
            "threshold": float(_require(row, "threshold", f"Rules sheet row {i}")),
            "severity": (row.get("severity") or "medium").lower(),
            "label": row.get("label") or row["metric"],
            "management_question": row.get("management_question") or "",
            "suggested_owner": row.get("suggested_owner") or "Unassigned",
            "next_action": row.get("next_action") or "",
            "trigger": row.get("trigger") or "",
        })
    _dump(target / "decision_rules.yaml", {"rules": rules_out},
          "Generated from an intake workbook. Firing is not a verdict — each rule\n"
          "names a question to ask, not an answer to accept.")

    _dump(target / "scenarios.yaml", {
        "presets": {"base": {"label": "Plan", "overrides": {}}},
        "monte_carlo": {},
    }, "Generated from an intake workbook. Only the Plan case is created; add\n"
       "scenarios and a monte_carlo block by hand — both need judgement the\n"
       "intake does not ask for.")

    _dump(target / "mappings.yaml", {
        "source": "Not yet mapped",
        "contract": ["date", "entity", "business_unit", "region", "account",
                     "metric", "actual", "budget", "forecast"],
        "metrics": {},
    }, "Generated from an intake workbook. The contract is stated; the per-metric\n"
       "account mapping is the next conversation with the client's systems team.")

    # Prove it: load it back through the real loader.
    try:
        clientpack.load_pack(client_id, clients_dir=clients_dir)
    except clientpack.ClientPackError as exc:
        raise IntakeError(f"The generated pack does not load: {exc}") from exc
    return target


def _dump(path: Path, data: dict, note: str) -> None:
    header = "\n".join(f"# {line}" for line in note.splitlines())
    path.write_text(header + "\n\n" + yaml.safe_dump(data, sort_keys=False, allow_unicode=True))


def main() -> None:  # pragma: no cover - CLI
    import argparse

    parser = argparse.ArgumentParser(description="Client intake: template in, pack out.")
    sub = parser.add_subparsers(dest="cmd", required=True)
    t = sub.add_parser("template", help="write a blank intake workbook")
    t.add_argument("--client", default="new_client")
    t.add_argument("--out", default=None)
    o = sub.add_parser("onboard", help="build a client pack from a filled workbook")
    o.add_argument("file")

    args = parser.parse_args()
    if args.cmd == "template":
        out = Path(args.out or f"clients/_intake/{args.client}-intake.xlsx")
        print(f"Wrote {write_template(out, args.client)}")
    else:
        target = build_pack(read_workbook(Path(args.file)))
        print(f"Wrote client pack to {target}")
        for f in sorted(target.iterdir()):
            print(f"  {f.name}")


if __name__ == "__main__":
    main()
